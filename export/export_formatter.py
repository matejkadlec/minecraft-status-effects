"""
Export formatter with theme-aware styling.
Handles CSV, XLSX, and JSON formatting.
"""

import json
import csv
import io
import re
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont


class ExportFormatter:
    # Theme colors from theme.css
    THEME_COLORS = {
        "light": {
            "header_bg": "2366b8",  # --header: #2366b8
            "header_text": "ffffff",  # white
            "row_bg_1": "ffffff",  # white
            "row_bg_2": "e6edf5",  # --bg-accent: #e6edf5
            "text": "000000",  # black
            "bold_text": "2366b8",  # blue for bold text
        },
        "dark": {
            "header_bg": "cfa93a",  # --header: #cfa93a
            "header_text": "000000",  # black
            "row_bg_1": "2d3035",  # --bg-accent: #2d3035 (grey)
            "row_bg_2": "24272b",  # --bg-alt: #24272b (darker grey)
            "text": "ffffff",  # white
            "bold_text": "cfa93a",  # gold for bold text
        },
    }

    # Column widths based on table.css (in Excel units, roughly)
    COLUMN_WIDTHS = {
        "mod": 25,  # 200px ≈ 25 Excel units
        "effect": 21,  # 170px ≈ 21 Excel units
        "maxLevel": 8,  # 50px ≈ 8 Excel units
        "description": 125,  # 1000px ≈ 125 Excel units
        "tags": 18,  # 145px ≈ 18 Excel units
        "source": 114,  # 910px ≈ 114 Excel units
    }

    def __init__(self, theme: str):
        """Initialize with theme (light or dark)."""
        self.theme = theme.lower()
        self.colors = self.THEME_COLORS[self.theme]

    def format_json(self, effects: List[Dict[str, Any]]) -> str:
        """Format effects as JSON."""
        return json.dumps({"effects": effects}, indent=2, ensure_ascii=False)

    def format_csv(self, effects: List[Dict[str, Any]]) -> str:
        """Format effects as CSV."""
        output = io.StringIO()

        headers = ["Mod", "Effect", "Max", "Description", "Tags", "Source"]
        writer = csv.writer(output)
        writer.writerow(headers)

        for effect in effects:
            # Strip HTML from description and source for CSV
            description = self._strip_html(effect["description"])
            source = self._strip_html(effect.get("source", ""))
            tags = ", ".join(effect["tags"])

            writer.writerow(
                [
                    effect["mod"],
                    effect["effect"],
                    effect["maxLevel"],
                    description,
                    tags,
                    source,
                ]
            )

        return output.getvalue()

    def format_xlsx(self, effects: List[Dict[str, Any]]) -> bytes:
        """Format effects as styled XLSX."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Status Effects"

        # Headers
        headers = ["Mod", "Effect", "Max", "Description", "Tags", "Source"]
        ws.append(headers)

        # Style header row
        self._style_header_row(ws, len(headers))

        # Add data rows
        for i, effect in enumerate(effects, start=2):
            # Convert tags list to string
            tags = ", ".join(effect["tags"])

            ws.append(
                [
                    effect["mod"],
                    effect["effect"],
                    effect["maxLevel"],
                    effect["description"],  # Keep HTML for styling
                    tags,
                    effect.get("source", ""),  # Keep HTML for styling
                ]
            )

            # Style data row
            self._style_data_row(ws, i, len(headers))

        # Set column widths
        self._set_column_widths(ws)

        # Add borders
        self._add_borders(ws, len(effects) + 1, len(headers))

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _style_header_row(self, ws, col_count: int):
        """Style the header row."""
        header_font = Font(
            name="Arial", bold=True, color=self.colors["header_text"], size=10.5
        )
        header_fill = PatternFill(
            start_color=self.colors["header_bg"],
            end_color=self.colors["header_bg"],
            fill_type="solid",
        )
        alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            # Convert to uppercase
            cell.value = cell.value.upper() if cell.value else ""

        # Set header row height
        ws.row_dimensions[1].height = 20

    def _style_data_row(self, ws, row_num: int, col_count: int):
        """Style a data row with alternating background."""
        # Determine row background (alternating)
        bg_color = (
            self.colors["row_bg_1"] if row_num % 2 == 0 else self.colors["row_bg_2"]
        )

        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        font = Font(name="Arial", color=self.colors["text"], size=10)

        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.font = font

            # Special handling for description and source columns (HTML formatting)
            if col in [4, 6]:  # Description and Source columns
                self._format_description_cell(cell)
                # Set alignment with wrap_text for description/source columns
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            # Center align Max Level and Tags columns
            elif col in [3, 5]:  # Max Level and Tags
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            # Left align for Mod and Effect columns
            else:  # Columns 1 and 2 (Mod and Effect)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

        # Set row height
        ws.row_dimensions[row_num].height = 18

    def _format_description_cell(self, cell):
        """Format description cell with HTML bold text styling."""
        if not cell.value:
            return

        text = str(cell.value)

        # Check if there are any bold tags
        if "<b>" not in text:
            # No bold formatting needed, just strip any other HTML
            cell.value = self._strip_html(text)
            return

        # Create rich text with bold formatting
        rich_text = CellRichText()

        # Split text by bold tags and process each part
        parts = re.split(r"(<b>.*?</b>)", text)

        # Regular font for normal text (using InlineFont for rich text)
        normal_font = InlineFont(rFont="Arial", color=self.colors["text"], sz=10)
        # Bold font with theme color for bold text
        bold_font = InlineFont(
            rFont="Arial", color=self.colors["bold_text"], sz=10, b=True
        )

        for part in parts:
            if part.startswith("<b>") and part.endswith("</b>"):
                # This is a bold section - remove tags and make it bold
                bold_text = part[3:-4]  # Remove <b> and </b>
                if bold_text:  # Only add if not empty
                    rich_text.append(TextBlock(bold_font, bold_text))
            elif part:  # Regular text (not empty)
                # Strip any remaining HTML tags and add as normal text
                clean_text = self._strip_html(part)
                if clean_text:  # Only add if not empty
                    rich_text.append(TextBlock(normal_font, clean_text))

        # Apply the rich text to the cell
        cell.value = rich_text

    def _set_column_widths(self, ws):
        """Set column widths based on our table.css values."""
        widths = [
            self.COLUMN_WIDTHS["mod"],
            self.COLUMN_WIDTHS["effect"],
            self.COLUMN_WIDTHS["maxLevel"],
            self.COLUMN_WIDTHS["description"],
            self.COLUMN_WIDTHS["tags"],
            self.COLUMN_WIDTHS["source"],
        ]

        for i, width in enumerate(widths, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

    def _add_borders(self, ws, row_count: int, col_count: int):
        """Add borders to the table."""
        thin_border = Side(style="thin", color="000000")
        medium_border = Side(style="medium", color="000000")

        # Header row borders
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.border = Border(
                left=thin_border if col > 1 else medium_border,
                right=medium_border if col == col_count else thin_border,
                top=medium_border,
                bottom=medium_border,
            )

        # Data row borders
        for row in range(2, row_count + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=thin_border if col > 1 else medium_border,
                    right=medium_border if col == col_count else thin_border,
                    top=thin_border if row > 2 else None,
                    bottom=medium_border if row == row_count else thin_border,
                )

    def _strip_html(self, text: str) -> str:
        """Remove HTML tags from text."""
        return re.sub(r"<[^>]+>", "", text)
